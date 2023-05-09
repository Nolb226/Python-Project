from pathlib import Path

from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage, ttk, END, messagebox
from tkcalendar import DateEntry
import tkinter.font as tkFont
import staff, checkin, checkout, shift
from tkinter.messagebox import askyesno
import openpyxl


class Schedule(ttk.Frame):
    def __init__(self,frame):
        super().__init__(frame)

        OUTPUT_PATH = Path(__file__).parent
        ASSETS_PATH = OUTPUT_PATH / Path(r".\assets\frame4")


        def relative_to_assets(path: str) -> Path:
            return ASSETS_PATH / Path(path)

        self.canvas = Canvas(
            self,
            bg = "#FFFFFF",
            height = 650,
            width = 650,
            bd = 0,
            highlightthickness = 0,
            relief = "ridge"
        )

        self.canvas.place(x = 0, y = 0)
        self.canvas.create_rectangle(
            431.0,
            48.0,
            1081.0,
            698.0,
            fill="#FFFFFF",
            outline="")

        self.canvas.create_text(
            280.0,
            65.0,
            anchor="nw",
            text="LỊCH LÀM",
            fill="#1E1E1E",
            font=("Inter SemiBold", 20 * -1)
        )

        self.canvas.create_text(
            100.0,
            105.0,
            anchor="nw",
            text="Từ",
            fill="#1E1E1E",
            font=("Inter SemiBold", 20 * -1)
        )
        self.start = DateEntry(
            self,
            state="readonly",
            date_pattern="yyyy-MM-dd"
        )
        self.start.place(
            x=140.0,
            y=105.0,
            width=80.0,
            height=30.0
        )

        self.canvas.create_text(
            240.0,
            105.0,
            anchor="nw",
            text="Đến",
            fill="#1E1E1E",
            font=("Inter SemiBold", 20 * -1)
        )
        self.end = DateEntry(
            self,
            state="readonly",
            date_pattern="yyyy-MM-dd"
        )
        self.end.place(
            x=290.0,
            y=105.0,
            width=80.0,
            height=30.0
        )

        self.button_image_4 = PhotoImage(
            file=relative_to_assets("button_4.png"))
        self.button_4 = Button(
            self,
            image=self.button_image_4,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: self.set_table(self.start.get(), self.end.get()),
            relief="flat"
        )
        self.button_4.place(
            x=390.0,
            y=105.0,
            width=65.0,
            height=30.0
        )

        self.button_image_5 = PhotoImage(
            file=relative_to_assets("button_5.png"))
        self.button_5 = Button(
            self,
            image=self.button_image_5,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: self.set_form(),
            relief="flat"
        )
        self.button_5.place(
            x=470.0,
            y=105.0,
            width=65.0,
            height=30.0
        )

        self.style = ttk.Style()
        self.style.configure("myStyle.Treeview", font=('Calibri', 10), rowheight=30) # Modify the font of the body
        self.style.configure("myStyle.Treeview.Heading", font=('Calibri', 10)) # Modify the font of the headings
        self.style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])

        self.scheduleTable = ttk.Treeview(self)

        self.entry_image_5 = PhotoImage(
            file=relative_to_assets("entry_1.png"))
        self.entry_bg_5 = self.canvas.create_image(
            300.0,
            170.0,
            image=self.entry_image_5
        )
        self.entry_5 = Entry(
            self,
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0
        )
        self.entry_5.place(
            x=104.0,
            y=156.0,
            width=310.0,
            height=26.0
        )
        self.button_image_3 = PhotoImage(
            file=relative_to_assets("button_8.png"))
        self.button_3 = Button(
            self,
            image=self.button_image_3,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: self.find_staff(self.entry_5.get()),
            relief="flat"
        )
        self.button_3.place(
            x=466.0,
            y=155.0,
            width=80.0,
            height=30.0
        )


        self.canvas.create_text(
            136.0,  # 567
            219.0,
            anchor="nw",
            text="Mã cá nhân",
            fill="#000000",
            font=("Inter SemiBold", 15 * -1)
        )

        self.entry_image_1 = PhotoImage(
            file=relative_to_assets("entry_2.png"))
        self.entry_bg_1 = self.canvas.create_image(
            386,
            229.0,
            image=self.entry_image_1
        )
        self.entry_1 = Entry(
            self,
            bd=0,
            bg="#FFFFFF",
            fg="#000716",
            highlightthickness=0,
        )
        self.entry_1.place(
            x=260,
            y=215.0,
            width=253.0,
            height=28.0
        )

        optionsH = ["0","1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23"]
        optionsM = ["00","15","30","45"]

        self.entry_image_2 = PhotoImage(
            file=relative_to_assets("entry_2.png"))
        self.entry_bg_2 = self.canvas.create_image(
            386,
            264.0,
            image=self.entry_image_2
        )
        self.entry_2 = ttk.Combobox(
            self,
            state="readonly",
            values=optionsH,
        )
        self.entry_2.place(
            x=260.0,
            y=250.0,
            width=126.0,
            height=28.0
        )
        self.entry_2_1 = ttk.Combobox(
            self,
            state="readonly",
            values=optionsM,
        )
        self.entry_2_1.place(
            x=386.0,
            y=250.0,
            width=126.0,
            height=28.0
        )

        self.canvas.create_text(
            136.0,
            254.0,
            anchor="nw",
            text="Từ",
            fill="#000000",
            font=("Inter SemiBold", 15 * -1)
        )

        
        self.entry_image_3 = PhotoImage(
            file=relative_to_assets("entry_2.png"))
        self.entry_bg_3 = self.canvas.create_image(
            386,
            299.0,
            image=self.entry_image_3
        )
        self.entry_3 = ttk.Combobox(
            self,
            state="readonly",
            values=optionsH,
        )
        self.entry_3.place(
            x=260,
            y=285.0,
            width=126.0,
            height=28.0
        )
        self.entry_3_1 = ttk.Combobox(
            self,
            state="readonly",
            values=optionsM,
        )
        self.entry_3_1.place(
            x=386,
            y=285.0,
            width=126.0,
            height=28.0
        )

        self.canvas.create_text(
            136.0,
            289.0,
            anchor="nw",
            text="Đến",
            fill="#000000",
            font=("Inter SemiBold", 15 * -1)
        )

        self.entry_image_4 = PhotoImage(
            file=relative_to_assets("entry_2.png"))
        self.entry_bg_4 = self.canvas.create_image(
            386,
            334.0,
            image=self.entry_image_3
        )
        self.entry_4 = DateEntry(
            self,
            state="readonly",
            date_pattern="yyyy-MM-dd"
        )
        self.entry_4.place(
            x=260,
            y=320.0,
            width=253.0,
            height=28.0
        )

        self.button_image_6 = PhotoImage(
            file=relative_to_assets("button_6.png"))
        self.button_6 = Button(
            self,
            image=self.button_image_6,
            borderwidth=0,
            highlightthickness=0,
            relief="flat"
        )
        self.button_6.place(
            x=290.0,
            y=369.0,
            width=70.0,
            height=30.0
        )
        self.canvas.create_text(
            136.0,
            324.0,
            anchor="nw",
            text="Ngày làm",
            fill="#000000",
            font=("Inter SemiBold", 15 * -1)
        )

        self.button_image_7 = PhotoImage(
            file=relative_to_assets("button_7.png"))
        self.button_7 = Button(
            self,
            image=self.button_image_7,
            borderwidth=0,
            highlightthickness=0,
            command=lambda: self.export(),
            relief="flat"
        )
        

        #right
        self.canvas.create_rectangle(
            527.0,
            199.0,
            529.0,
            409.0,
            fill="#000000",
            outline="")

        #bot
        self.canvas.create_rectangle(
            123.0,
            408.0,
            527.0,
            410.0,
            fill="#000000",
            outline="")

        #left
        self.canvas.create_rectangle(
            123.0,
            199.0,
            125.0,
            408.0,
            fill="#000000",
            outline="")

        #top
        self.canvas.create_rectangle(
            123.0,
            198.0,
            527.0,
            200.0,
            fill="#000000",
            outline="")
        

    def set_table(self, start, end):
        self.scheduleTable.place_forget()
        day = shift.select_day(start, end)
        id = staff.select_id()
        self.columns = ["ID", "FULL NAME"]
        self.rows = []
        for i in range(len(day)):
            self.columns.append(day[i][0].strftime("%Y-%m-%d"))
        self.scheduleTable = ttk.Treeview(self, columns=tuple(self.columns), show="headings", style="myStyle.Treeview")
        for i in range(len(self.columns)):
            self.scheduleTable.heading(i, text=self.columns[i])
            if i == 0:
                self.scheduleTable.column(i, width=70, anchor='center')
            else:
                self.scheduleTable.column(i, width=150, anchor='center')
        self.set_data_table(self.columns,id)
        self.vsbx = ttk.Scrollbar(self, orient="horizontal", command=self.scheduleTable.xview)
        self.vsbx.place(x=0,y=595,width=610)
        self.vsby = ttk.Scrollbar(self, orient="vertical", command=self.scheduleTable.yview)
        self.vsby.place(x=622,y=150,height=463)
        self.scheduleTable.configure(yscrollcommand=self.vsby.set, xscrollcommand=self.vsbx.set)
        self.scheduleTable.place(x=0,y=150, width=610,height=440)
        self.button_7.place(
            x=250.0,
            y=620.0,
            width=130.0,
            height=30.0
        )

    def set_data_table(self,day,id):
        for i in id:
            row = [i, staff.select_one_name(i)]
            for d in day:
                if d != "ID" and d != "FULL NAME":
                    row.append(shift.select_shift(i,d))
            self.scheduleTable.insert('','end',values=row)

    def set_form(self):
        self.scheduleTable.place_forget()
        self.vsbx.place_forget()
        self.vsby.place_forget()
        self.button_7.place_forget()

        # =================================================
        

    def find_staff(self,input):
        if input in staff.select_id():
            self.entry_1.delete(0, END)
            self.entry_1.insert(0, input)
            self.button_6.config(command=lambda: self.insert())
        else:
            messagebox.showerror("Lỗi","Không tìm thấy nhân viên")
            self.entry_5.delete(0, END)

    def insert(self):
        answer = askyesno(title="Confirm insert",
                        message='Bạn chắc chắn chứ ?')

        if answer:
            if self.entry_1.get() == '' or self.entry_2.get() == '' or self.entry_4.get() == '':

                messagebox.showerror(
                    'Invalid input', 'Thông tin không được để trống')
                return
            if not self.entry_1.get().isdigit():
                messagebox.showerror(
                    'Invalid input', 'Mã cá nhân phải là số')
                return
            try:
                inTime = "{}:{}:00".format(self.entry_2.get(),self.entry_2_1.get())
                outTime = "{}:{}:00".format(self.entry_3.get(),self.entry_3_1.get())
                a = shift.Shift(self.entry_1.get(),self.entry_4.get(),inTime=inTime,outTime=outTime)
                a.insert()
            except ValueError:
                messagebox.showerror(
                    'ERROR ⚠', 'Mã trùng')
                return
            messagebox.showinfo('Done!!!', "Đã thêm thành công")

            self.entry_1.delete(0, END)
            self.entry_2.delete(0, END)

    def export(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # Ghi heading vào file Excel
        for i in range(len(self.columns)):
            worksheet.cell(row=1,column=i+1,value=self.columns[i])
        for i, row in enumerate(self.scheduleTable.get_children()):
            values = self.scheduleTable.item(row)["values"]
            for j in range(len(values)):
                worksheet.cell(row=i+2,column=j+1,value=values[j])
        workbook.save("{}-{}.xlsx".format(self.columns[2],self.columns[len(self.columns)-1]))


